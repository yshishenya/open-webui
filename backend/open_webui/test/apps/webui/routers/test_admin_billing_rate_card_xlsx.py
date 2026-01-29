import io
import time
import uuid

import openpyxl

from test.util.abstract_integration_test import AbstractPostgresTest
from test.util.mock_user import mock_webui_user


class TestAdminBillingRateCardXlsx(AbstractPostgresTest):
    BASE_PATH = "/api/v1/admin/billing"

    def setup_method(self) -> None:
        super().setup_method()

        from open_webui.models.billing import PricingRateCardModel, RateCards
        from open_webui.models.models import ModelForm, ModelMeta, ModelParams, Models

        self.model_a = "rate-card-model-a"
        self.model_b = "rate-card-model-b"

        # Create base models so model IDs are known to the system.
        Models.insert_new_model(
            ModelForm(
                id=self.model_a,
                base_model_id=None,
                name="Model A",
                meta=ModelMeta(),
                params=ModelParams(),
                access_control=None,
                is_active=True,
            ),
            user_id="admin",
        )
        Models.insert_new_model(
            ModelForm(
                id=self.model_b,
                base_model_id=None,
                name="Model B",
                meta=ModelMeta(),
                params=ModelParams(),
                access_control=None,
                is_active=True,
            ),
            user_id="admin",
        )

        now = int(time.time())

        self.rate_a_token_in_id = str(uuid.uuid4())
        RateCards.create_rate_card(
            PricingRateCardModel(
                id=self.rate_a_token_in_id,
                model_id=self.model_a,
                model_tier=None,
                modality="text",
                unit="token_in",
                raw_cost_per_unit_kopeks=100,
                version="2025-01",
                created_at=now,
                provider="Test",
                is_default=True,
                is_active=True,
            ).model_dump()
        )

        self.rate_a_token_out_id = str(uuid.uuid4())
        RateCards.create_rate_card(
            PricingRateCardModel(
                id=self.rate_a_token_out_id,
                model_id=self.model_a,
                model_tier=None,
                modality="text",
                unit="token_out",
                raw_cost_per_unit_kopeks=110,
                version="2025-01",
                created_at=now,
                provider="Test",
                is_default=False,
                is_active=True,
            ).model_dump()
        )

        self.rate_b_image_id = str(uuid.uuid4())
        RateCards.create_rate_card(
            PricingRateCardModel(
                id=self.rate_b_image_id,
                model_id=self.model_b,
                model_tier=None,
                modality="image",
                unit="image_1024",
                raw_cost_per_unit_kopeks=250,
                version="2025-01",
                created_at=now,
                provider="Test",
                is_default=True,
                is_active=True,
            ).model_dump()
        )

    def _read_workbook(self, content: bytes) -> openpyxl.Workbook:
        return openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    def test_export_xlsx_active_only(self, monkeypatch) -> None:
        from open_webui.routers import admin_billing_rate_card

        monkeypatch.setattr(admin_billing_rate_card, "ENABLE_BILLING_WALLET", True)
        monkeypatch.setattr(admin_billing_rate_card, "BILLING_RATE_CARD_VERSION", "2025-01")

        with mock_webui_user(role="admin"):
            response = self.fast_api_client.get(
                self.create_url(
                    "/rate-card/export-xlsx",
                    query_params={
                        "model_ids": self.model_a,
                        "mode": "active_only",
                    },
                )
            )

        assert response.status_code == 200
        assert (
            response.headers.get("content-type")
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        wb = self._read_workbook(response.content)
        assert "RateCards" in wb.sheetnames

        ws = wb["RateCards"]
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        assert "model_id" in header
        assert "raw_cost_per_unit_kopeks" in header

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        # active_only exports only active units for the selected models.
        assert len(rows) == 2
        assert set(r[0] for r in rows) == {self.model_a}

    def test_export_xlsx_all_units_template(self, monkeypatch) -> None:
        from open_webui.routers import admin_billing_rate_card

        monkeypatch.setattr(admin_billing_rate_card, "ENABLE_BILLING_WALLET", True)
        monkeypatch.setattr(admin_billing_rate_card, "BILLING_RATE_CARD_VERSION", "2025-01")

        with mock_webui_user(role="admin"):
            response = self.fast_api_client.get(
                self.create_url(
                    "/rate-card/export-xlsx",
                    query_params={
                        "model_ids": self.model_a,
                        "mode": "all_units_template",
                    },
                )
            )

        assert response.status_code == 200
        wb = self._read_workbook(response.content)
        ws = wb["RateCards"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        # Template has exactly 5 rows per model.
        assert len(rows) == 5
        assert set(r[0] for r in rows) == {self.model_a}

    def test_import_preview_warns_out_of_scope_and_unknown(self, monkeypatch) -> None:
        from open_webui.routers import admin_billing_rate_card
        from open_webui.utils.rate_card_xlsx import dump_scope_model_ids

        monkeypatch.setattr(admin_billing_rate_card, "ENABLE_BILLING_WALLET", True)
        monkeypatch.setattr(admin_billing_rate_card, "BILLING_RATE_CARD_VERSION", "2025-01")

        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "RateCards"
        ws.append(
            [
                "model_id",
                "modality",
                "unit",
                "is_active",
                "raw_cost_per_unit_kopeks",
            ]
        )
        # Unknown model (exists neither in system nor in scope)
        ws.append(["unknown-model", "text", "token_in", True, 10])
        # Existing model but not in scope
        ws.append([self.model_b, "image", "image_1024", True, 123])

        buf = io.BytesIO()
        wb.save(buf)

        with mock_webui_user(role="admin"):
            response = self.fast_api_client.post(
                self.create_url("/rate-card/import-xlsx/preview"),
                files={
                    "file": (
                        "rate-cards.xlsx",
                        buf.getvalue(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
                data={
                    "mode": "patch",
                    "scope_model_ids": dump_scope_model_ids([self.model_a]),
                },
            )

        assert response.status_code == 200
        payload = response.json()
        warnings = payload.get("warnings", [])
        assert any(w.get("code") == "unknown_model" for w in warnings)
        assert any(w.get("code") == "model_out_of_scope" for w in warnings)

    def test_import_apply_updates_price_and_deactivates_old(self, monkeypatch) -> None:
        from open_webui.models.billing import RateCards
        from open_webui.routers import admin_billing_rate_card
        from open_webui.utils.rate_card_xlsx import dump_scope_model_ids

        monkeypatch.setattr(admin_billing_rate_card, "ENABLE_BILLING_WALLET", True)
        monkeypatch.setattr(admin_billing_rate_card, "BILLING_RATE_CARD_VERSION", "2025-01")

        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "RateCards"
        ws.append(
            [
                "model_id",
                "modality",
                "unit",
                "is_active",
                "raw_cost_per_unit_kopeks",
                "provider",
                "is_default",
            ]
        )
        # Update token_in price from 100 to 150.
        ws.append([self.model_a, "text", "token_in", True, 150.0, "ProviderX", True])

        buf = io.BytesIO()
        wb.save(buf)

        with mock_webui_user(role="admin"):
            response = self.fast_api_client.post(
                self.create_url("/rate-card/import-xlsx/apply"),
                files={
                    "file": (
                        "rate-cards.xlsx",
                        buf.getvalue(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
                data={
                    "mode": "patch",
                    "scope_model_ids": dump_scope_model_ids([self.model_a]),
                },
            )

        assert response.status_code == 200
        payload = response.json()
        assert payload.get("errors") == []

        # Old should now be inactive.
        old = RateCards.get_rate_card_by_id(self.rate_a_token_in_id)
        assert old is not None
        assert old.is_active is False

        # New active should exist.
        active = RateCards.get_active_rate_card(self.model_a, "text", "token_in")
        assert active is not None
        assert active.raw_cost_per_unit_kopeks == 150
        assert active.provider == "ProviderX"
        assert active.is_default is True

    def test_import_preview_missing_price_error_when_active(self, monkeypatch) -> None:
        from open_webui.routers import admin_billing_rate_card
        from open_webui.utils.rate_card_xlsx import dump_scope_model_ids

        monkeypatch.setattr(admin_billing_rate_card, "ENABLE_BILLING_WALLET", True)
        monkeypatch.setattr(admin_billing_rate_card, "BILLING_RATE_CARD_VERSION", "2025-01")

        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "RateCards"
        ws.append(["model_id", "modality", "unit", "is_active", "raw_cost_per_unit_kopeks"])
        ws.append([self.model_a, "text", "token_in", True, ""])

        buf = io.BytesIO()
        wb.save(buf)

        with mock_webui_user(role="admin"):
            response = self.fast_api_client.post(
                self.create_url("/rate-card/import-xlsx/preview"),
                files={
                    "file": (
                        "rate-cards.xlsx",
                        buf.getvalue(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
                data={
                    "mode": "patch",
                    "scope_model_ids": dump_scope_model_ids([self.model_a]),
                },
            )

        assert response.status_code == 200
        payload = response.json()
        errors = payload.get("errors", [])
        assert any(e.get("code") == "missing_price" for e in errors)

    def test_import_apply_returns_400_with_structured_errors(self, monkeypatch) -> None:
        from open_webui.routers import admin_billing_rate_card
        from open_webui.utils.rate_card_xlsx import dump_scope_model_ids

        monkeypatch.setattr(admin_billing_rate_card, "ENABLE_BILLING_WALLET", True)
        monkeypatch.setattr(admin_billing_rate_card, "BILLING_RATE_CARD_VERSION", "2025-01")

        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "RateCards"
        ws.append(["model_id", "modality", "unit", "is_active", "raw_cost_per_unit_kopeks"])
        # Invalid: active row but missing price.
        ws.append([self.model_a, "text", "token_in", True, ""])

        buf = io.BytesIO()
        wb.save(buf)

        with mock_webui_user(role="admin"):
            response = self.fast_api_client.post(
                self.create_url("/rate-card/import-xlsx/apply"),
                files={
                    "file": (
                        "rate-cards.xlsx",
                        buf.getvalue(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
                data={
                    "mode": "patch",
                    "scope_model_ids": dump_scope_model_ids([self.model_a]),
                },
            )

        assert response.status_code == 400
        payload = response.json()
        errors = payload.get("errors", [])
        assert any(e.get("code") == "missing_price" for e in errors)

    def test_import_preview_rejects_non_integer_float_price(self, monkeypatch) -> None:
        from open_webui.routers import admin_billing_rate_card
        from open_webui.utils.rate_card_xlsx import dump_scope_model_ids

        monkeypatch.setattr(admin_billing_rate_card, "ENABLE_BILLING_WALLET", True)
        monkeypatch.setattr(admin_billing_rate_card, "BILLING_RATE_CARD_VERSION", "2025-01")

        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "RateCards"
        ws.append(["model_id", "modality", "unit", "is_active", "raw_cost_per_unit_kopeks"])
        ws.append([self.model_a, "text", "token_in", True, 150.5])

        buf = io.BytesIO()
        wb.save(buf)

        with mock_webui_user(role="admin"):
            response = self.fast_api_client.post(
                self.create_url("/rate-card/import-xlsx/preview"),
                files={
                    "file": (
                        "rate-cards.xlsx",
                        buf.getvalue(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
                data={
                    "mode": "patch",
                    "scope_model_ids": dump_scope_model_ids([self.model_a]),
                },
            )

        assert response.status_code == 200
        payload = response.json()
        errors = payload.get("errors", [])
        assert any(e.get("code") == "invalid_price" for e in errors)

    def test_import_apply_full_sync_deactivates_missing_units(self, monkeypatch) -> None:
        from open_webui.models.billing import RateCards
        from open_webui.routers import admin_billing_rate_card
        from open_webui.utils.rate_card_xlsx import dump_scope_model_ids

        monkeypatch.setattr(admin_billing_rate_card, "ENABLE_BILLING_WALLET", True)
        monkeypatch.setattr(admin_billing_rate_card, "BILLING_RATE_CARD_VERSION", "2025-01")

        # Ensure model_a token_out is active initially.
        assert RateCards.get_active_rate_card(self.model_a, "text", "token_out") is not None

        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "RateCards"
        ws.append(["model_id", "modality", "unit", "is_active", "raw_cost_per_unit_kopeks"])
        # Only include token_in, token_out should be deactivated by full_sync.
        ws.append([self.model_a, "text", "token_in", True, 100])

        buf = io.BytesIO()
        wb.save(buf)

        with mock_webui_user(role="admin"):
            response = self.fast_api_client.post(
                self.create_url("/rate-card/import-xlsx/apply"),
                files={
                    "file": (
                        "rate-cards.xlsx",
                        buf.getvalue(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
                data={
                    "mode": "full_sync",
                    "scope_model_ids": dump_scope_model_ids([self.model_a]),
                },
            )

        assert response.status_code == 200
        payload = response.json()
        assert payload.get("errors") == []
        assert RateCards.get_active_rate_card(self.model_a, "text", "token_out") is None
