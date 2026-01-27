import io
import json
import logging
from dataclasses import dataclass
from enum import Enum
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook

from open_webui.models.billing import PricingRateCardModel

log = logging.getLogger(__name__)


class XlsxExportMode(str, Enum):
    ACTIVE_ONLY = "active_only"
    ALL_UNITS_TEMPLATE = "all_units_template"


class XlsxImportMode(str, Enum):
    PATCH = "patch"
    FULL_SYNC = "full_sync"


@dataclass(frozen=True)
class RateCardUnit:
    modality: str
    unit: str


ALLOWED_UNITS: Tuple[RateCardUnit, ...] = (
    RateCardUnit(modality="text", unit="token_in"),
    RateCardUnit(modality="text", unit="token_out"),
    RateCardUnit(modality="image", unit="image_1024"),
    RateCardUnit(modality="tts", unit="tts_char"),
    RateCardUnit(modality="stt", unit="stt_second"),
)

ALLOWED_MODALITIES = {"text", "image", "tts", "stt"}

ALLOWED_UNITS_BY_MODALITY: Dict[str, set[str]] = {
    "text": {"token_in", "token_out"},
    "image": {"image_1024"},
    "tts": {"tts_char"},
    "stt": {"stt_second"},
}


@dataclass
class XlsxWarning:
    row_number: int
    code: str
    message: str
    model_id: Optional[str] = None


@dataclass
class XlsxError:
    code: str
    message: str
    row_number: Optional[int] = None
    column: Optional[str] = None


@dataclass
class XlsxImportSummary:
    rows_total: int
    rows_valid: int
    rows_invalid: int
    creates: int
    updates_via_create: int
    deactivations: int
    noops: int
    skipped_unknown_model: int
    skipped_out_of_scope: int


@dataclass(frozen=True)
class ParsedRateCardRow:
    row_number: int
    model_id: str
    modality: str
    unit: str
    is_active: bool
    raw_cost_per_unit_kopeks: Optional[int]
    provider: Optional[str]
    model_tier: Optional[str]
    is_default: Optional[bool]


@dataclass(frozen=True)
class ImportAction:
    action: str
    model_id: str
    modality: str
    unit: str
    desired_active: bool
    desired_price: Optional[int]
    provider: Optional[str]
    model_tier: Optional[str]
    is_default: Optional[bool]


def normalize_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_empty(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def parse_bool(value: object) -> Optional[bool]:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        if value == 1:
            return True
        if value == 0:
            return False
    text = normalize_str(value).lower()
    if text == "":
        return None
    if text in {"true", "1", "yes"}:
        return True
    if text in {"false", "0", "no"}:
        return False
    return None


def parse_price_kopeks(value: object) -> Tuple[Optional[int], Optional[str]]:
    """Return (parsed_int, error_message)."""

    if value is None:
        return None, None

    if isinstance(value, bool):
        return None, "Price must be a number"

    if isinstance(value, int):
        if value < 0:
            return None, "Price must be >= 0"
        return value, None

    if isinstance(value, float):
        if value < 0:
            return None, "Price must be >= 0"
        if value.is_integer():
            return int(value), None
        return None, "Price must be an integer (or *.0)"

    text = normalize_str(value)
    if text == "":
        return None, None

    try:
        if "." in text:
            parsed_float = float(text)
            if parsed_float < 0:
                return None, "Price must be >= 0"
            if parsed_float.is_integer():
                return int(parsed_float), None
            return None, "Price must be an integer (or *.0)"
        parsed_int = int(text)
        if parsed_int < 0:
            return None, "Price must be >= 0"
        return parsed_int, None
    except Exception:
        return None, "Price must be a number"


def dump_scope_model_ids(scope_model_ids: Sequence[str]) -> str:
    return json.dumps(list(scope_model_ids))


def parse_scope_model_ids(value: str) -> Tuple[Optional[List[str]], Optional[str]]:
    """Return (ids, error)."""

    text = normalize_str(value)
    if not text:
        return None, "scope_model_ids is required"

    try:
        parsed = json.loads(text)
    except Exception:
        return None, "scope_model_ids must be a JSON array"

    if not isinstance(parsed, list):
        return None, "scope_model_ids must be a JSON array"

    ids: List[str] = []
    for item in parsed:
        item_text = normalize_str(item)
        if item_text:
            ids.append(item_text)

    if not ids:
        return None, "scope_model_ids must contain at least one model id"

    return ids, None


def build_rate_card_key(model_id: str, modality: str, unit: str) -> str:
    return f"{model_id}:{modality}:{unit}"


def build_latest_active_index(
    entries: Iterable[PricingRateCardModel],
) -> Dict[str, PricingRateCardModel]:
    """Defensive: if multiple active rows exist, pick max created_at."""

    index: Dict[str, PricingRateCardModel] = {}
    for entry in entries:
        if not entry.is_active:
            continue
        key = build_rate_card_key(entry.model_id, entry.modality, entry.unit)
        current = index.get(key)
        if current is None or entry.created_at >= current.created_at:
            index[key] = entry
    return index


def build_export_workbook(
    *,
    mode: XlsxExportMode,
    model_ids: Sequence[str],
    model_names: Dict[str, str],
    active_entries: Iterable[PricingRateCardModel],
) -> bytes:
    """Create XLSX bytes for export.

    active_entries must already be filtered to BILLING_RATE_CARD_VERSION.
    """

    active_index = build_latest_active_index(active_entries)

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Failed to create workbook")
    ws.title = "RateCards"

    headers = [
        "model_id",
        "model_name",
        "modality",
        "unit",
        "is_active",
        "raw_cost_per_unit_kopeks",
        "provider",
        "model_tier",
        "is_default",
        "comment",
    ]
    ws.append(headers)

    if mode == XlsxExportMode.ACTIVE_ONLY:
        for entry in sorted(
            active_index.values(),
            key=lambda e: (e.model_id, e.modality, e.unit),
        ):
            ws.append(
                [
                    entry.model_id,
                    model_names.get(entry.model_id, ""),
                    entry.modality,
                    entry.unit,
                    True,
                    entry.raw_cost_per_unit_kopeks,
                    entry.provider or "",
                    entry.model_tier or "",
                    bool(entry.is_default),
                    "",
                ]
            )

    elif mode == XlsxExportMode.ALL_UNITS_TEMPLATE:
        for model_id in model_ids:
            for allowed in ALLOWED_UNITS:
                key = build_rate_card_key(model_id, allowed.modality, allowed.unit)
                active = active_index.get(key)
                if active:
                    ws.append(
                        [
                            model_id,
                            model_names.get(model_id, ""),
                            allowed.modality,
                            allowed.unit,
                            True,
                            active.raw_cost_per_unit_kopeks,
                            active.provider or "",
                            active.model_tier or "",
                            bool(active.is_default),
                            "",
                        ]
                    )
                else:
                    ws.append(
                        [
                            model_id,
                            model_names.get(model_id, ""),
                            allowed.modality,
                            allowed.unit,
                            False,
                            "",
                            "",
                            "",
                            "",
                            "",
                        ]
                    )
    else:
        raise ValueError(f"Unsupported export mode: {mode}")

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def parse_import_workbook(file_bytes: bytes) -> Tuple[List[ParsedRateCardRow], List[XlsxError]]:
    errors: List[XlsxError] = []

    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception:
        return [], [XlsxError(code="invalid_template", message="Invalid XLSX file")]

    if "RateCards" not in wb.sheetnames:
        return [], [
            XlsxError(code="invalid_template", message="Missing required sheet: RateCards")
        ]

    ws = wb["RateCards"]
    assert ws is not None

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], [XlsxError(code="invalid_template", message="Empty workbook")]

    header_row = rows[0]
    header_map: Dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        name = normalize_str(cell)
        if name:
            header_map[name] = idx

    required_headers = [
        "model_id",
        "modality",
        "unit",
        "raw_cost_per_unit_kopeks",
    ]

    missing = [h for h in required_headers if h not in header_map]
    if missing:
        return [], [
            XlsxError(
                code="invalid_template",
                message=f"Missing required columns: {', '.join(missing)}",
            )
        ]

    parsed_rows: List[ParsedRateCardRow] = []

    for excel_idx, row in enumerate(rows[1:], start=2):
        model_id = normalize_str(row[header_map["model_id"]] if header_map.get("model_id") is not None else "")
        modality = normalize_str(row[header_map["modality"]]).lower()
        unit = normalize_str(row[header_map["unit"]]).lower()

        is_active_val = None
        if "is_active" in header_map:
            is_active_val = parse_bool(row[header_map["is_active"]])
        is_active = True if is_active_val is None else bool(is_active_val)

        price_cell = row[header_map["raw_cost_per_unit_kopeks"]]
        price, price_error = parse_price_kopeks(price_cell)

        provider: Optional[str] = None
        if "provider" in header_map:
            provider_raw = row[header_map["provider"]]
            if not _is_empty(provider_raw):
                provider = normalize_str(provider_raw)

        model_tier: Optional[str] = None
        if "model_tier" in header_map:
            model_tier_raw = row[header_map["model_tier"]]
            if not _is_empty(model_tier_raw):
                model_tier = normalize_str(model_tier_raw)

        is_default: Optional[bool] = None
        if "is_default" in header_map:
            is_default_raw = row[header_map["is_default"]]
            is_default = parse_bool(is_default_raw)

        parsed_rows.append(
            ParsedRateCardRow(
                row_number=excel_idx,
                model_id=model_id,
                modality=modality,
                unit=unit,
                is_active=is_active,
                raw_cost_per_unit_kopeks=price,
                provider=provider,
                model_tier=model_tier,
                is_default=is_default,
            )
        )

        if not model_id:
            errors.append(
                XlsxError(
                    code="missing_model_id",
                    message="model_id is required",
                    row_number=excel_idx,
                    column="model_id",
                )
            )

        if modality not in ALLOWED_MODALITIES:
            errors.append(
                XlsxError(
                    code="invalid_modality",
                    message="Invalid modality",
                    row_number=excel_idx,
                    column="modality",
                )
            )
        else:
            allowed_units = ALLOWED_UNITS_BY_MODALITY.get(modality, set())
            if unit not in allowed_units:
                errors.append(
                    XlsxError(
                        code="invalid_unit",
                        message="Invalid unit for modality",
                        row_number=excel_idx,
                        column="unit",
                    )
                )

        if price_error:
            errors.append(
                XlsxError(
                    code="invalid_price",
                    message=price_error,
                    row_number=excel_idx,
                    column="raw_cost_per_unit_kopeks",
                )
            )

        if is_active and price is None:
            errors.append(
                XlsxError(
                    code="missing_price",
                    message="raw_cost_per_unit_kopeks is required when is_active is true",
                    row_number=excel_idx,
                    column="raw_cost_per_unit_kopeks",
                )
            )

    return parsed_rows, errors


def compute_import_plan(
    *,
    rows: Sequence[ParsedRateCardRow],
    import_mode: XlsxImportMode,
    scope_model_ids: Sequence[str],
    known_model_ids: set[str],
    current_active_entries: Iterable[PricingRateCardModel],
    base_errors: Sequence[XlsxError] = (),
) -> Tuple[XlsxImportSummary, List[XlsxWarning], List[XlsxError], List[ImportAction]]:
    warnings: List[XlsxWarning] = []
    errors: List[XlsxError] = list(base_errors)

    invalid_row_numbers = {
        err.row_number for err in base_errors if err.row_number is not None
    }

    scope_set = set(scope_model_ids)

    active_index = build_latest_active_index(current_active_entries)

    seen_keys: Dict[str, int] = {}

    creates = 0
    updates_via_create = 0
    deactivations = 0
    noops = 0
    skipped_unknown_model = 0
    skipped_out_of_scope = 0

    actions: List[ImportAction] = []

    valid_rows: List[ParsedRateCardRow] = []

    for row in rows:
        if row.row_number in invalid_row_numbers:
            continue
        if not row.model_id or not row.modality or not row.unit:
            continue

        key = build_rate_card_key(row.model_id, row.modality, row.unit)

        if key in seen_keys:
            errors.append(
                XlsxError(
                    code="duplicate_key",
                    message=f"Duplicate key also found at row {seen_keys[key]}",
                    row_number=row.row_number,
                )
            )
            invalid_row_numbers.add(row.row_number)
            continue
        seen_keys[key] = row.row_number

        if row.model_id not in known_model_ids:
            skipped_unknown_model += 1
            warnings.append(
                XlsxWarning(
                    row_number=row.row_number,
                    code="unknown_model",
                    message="Model does not exist in the system; row skipped",
                    model_id=row.model_id,
                )
            )
            continue

        if row.model_id not in scope_set:
            skipped_out_of_scope += 1
            warnings.append(
                XlsxWarning(
                    row_number=row.row_number,
                    code="model_out_of_scope",
                    message="Model is not selected (out of scope); row skipped",
                    model_id=row.model_id,
                )
            )
            continue

        valid_rows.append(row)

        existing_active = active_index.get(key)
        if row.is_active:
            if existing_active is None:
                creates += 1
                actions.append(
                    ImportAction(
                        action="create",
                        model_id=row.model_id,
                        modality=row.modality,
                        unit=row.unit,
                        desired_active=True,
                        desired_price=row.raw_cost_per_unit_kopeks,
                        provider=row.provider,
                        model_tier=row.model_tier,
                        is_default=row.is_default,
                    )
                )
            else:
                desired_price = int(row.raw_cost_per_unit_kopeks or 0)
                if desired_price == int(existing_active.raw_cost_per_unit_kopeks or 0):
                    noops += 1
                    actions.append(
                        ImportAction(
                            action="noop",
                            model_id=row.model_id,
                            modality=row.modality,
                            unit=row.unit,
                            desired_active=True,
                            desired_price=desired_price,
                            provider=None,
                            model_tier=None,
                            is_default=None,
                        )
                    )
                else:
                    updates_via_create += 1
                    actions.append(
                        ImportAction(
                            action="update_via_create",
                            model_id=row.model_id,
                            modality=row.modality,
                            unit=row.unit,
                            desired_active=True,
                            desired_price=desired_price,
                            provider=row.provider,
                            model_tier=row.model_tier,
                            is_default=row.is_default,
                        )
                    )
        else:
            if existing_active is None:
                noops += 1
                actions.append(
                    ImportAction(
                        action="noop",
                        model_id=row.model_id,
                        modality=row.modality,
                        unit=row.unit,
                        desired_active=False,
                        desired_price=None,
                        provider=None,
                        model_tier=None,
                        is_default=None,
                    )
                )
            else:
                deactivations += 1
                actions.append(
                    ImportAction(
                        action="deactivate",
                        model_id=row.model_id,
                        modality=row.modality,
                        unit=row.unit,
                        desired_active=False,
                        desired_price=None,
                        provider=None,
                        model_tier=None,
                        is_default=None,
                    )
                )

    if import_mode == XlsxImportMode.FULL_SYNC:
        present_models = {row.model_id for row in valid_rows}
        present_keys = {build_rate_card_key(r.model_id, r.modality, r.unit) for r in valid_rows}

        for model_id in present_models:
            for allowed in ALLOWED_UNITS:
                key = build_rate_card_key(model_id, allowed.modality, allowed.unit)
                if key in present_keys:
                    continue
                existing_active = active_index.get(key)
                if existing_active is None:
                    continue
                deactivations += 1
                actions.append(
                    ImportAction(
                        action="deactivate",
                        model_id=model_id,
                        modality=allowed.modality,
                        unit=allowed.unit,
                        desired_active=False,
                        desired_price=None,
                        provider=None,
                        model_tier=None,
                        is_default=None,
                    )
                )

    rows_total = len(rows)
    rows_invalid = len(invalid_row_numbers)
    rows_valid_count = max(0, rows_total - rows_invalid)

    summary = XlsxImportSummary(
        rows_total=rows_total,
        rows_valid=rows_valid_count,
        rows_invalid=rows_invalid,
        creates=creates,
        updates_via_create=updates_via_create,
        deactivations=deactivations,
        noops=noops,
        skipped_unknown_model=skipped_unknown_model,
        skipped_out_of_scope=skipped_out_of_scope,
    )

    return summary, warnings, errors, actions
